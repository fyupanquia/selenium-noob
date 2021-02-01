import unittest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
import time


class Excel(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome(
            executable_path=r"../../chromedriver.exe")
        self.driver.maximize_window()

    def test_excel(self):
        try:
            driver = self.driver
            driver.get("https://google.com")
            file_path = "./data.xlsx"
            wb = load_workbook(file_path)
            sheet_names = wb.get_sheet_names();
            print("**"*20)
            print(sheet_names)
            sheet = wb.get_sheet_by_name(sheet_names[0])
            wb.close()
            print("MAX_ROW: "+str(sheet.max_row))
            for i in range(2, sheet.max_row):
                name, age = sheet[f'A{i}:B{i}'][0]
                print("NAME: "+str(name.value)+" AGE:"+str(age.value))
            time.sleep(5)
            print("**"*20)

        except NoSuchElementException as ex:
            self.fail(ex.msg)

    def tearDow(self):
        """Stop web driver"""
        self.driver.quit()


if __name__ == "__main__":
    unittest.main(
        verbosity=2)
